import datetime
from io import BytesIO

import frappe
import openpyxl
from frappe.desk.query_report import run
from frappe.utils.xlsxutils import handle_html
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

@frappe.whitelist()
def export_query(report_name, filters=None):
    """
    Custom Global API to export Frappe Query Reports as styled Excel.
    """
    if filters:
        filters = frappe.parse_json(filters)
    else:
        filters = {}

    # =========================
    # FETCH REPORT DATA
    # =========================
    report_data = run(
        report_name,
        filters=filters,
        ignore_prepared_report=True,
    )

    columns = report_data.get("columns", [])
    data = report_data.get("result", [])

    # =========================
    # WORKBOOK & STYLES
    # =========================
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    ws = wb.create_sheet(title="Report")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"), right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"), bottom=Side(style="thin", color="D3D3D3"),
    )
    alternate_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    # =========================
    # HEADER
    # =========================
    for col_idx, column in enumerate(columns, start=1):
        label = column.get("label", "")
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # =========================
    # DATA ROWS
    # =========================
    for row_idx, row in enumerate(data, start=2):
        for col_idx, column in enumerate(columns, start=1):
            fieldname = column.get("fieldname")
            label = column.get("label")

            if isinstance(row, dict):
                value = row.get(fieldname, row.get(label, ""))
            else:
                value = row[col_idx - 1] if len(row) >= col_idx else ""

            if isinstance(value, str):
                value = handle_html(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Zebra rows
            if row_idx % 2 == 0:
                cell.fill = alternate_fill

            # Date formatting
            if isinstance(value, (datetime.date, datetime.datetime)):
                number_format = "DD-MM-YYYY"
                if isinstance(value, datetime.datetime):
                    number_format = "DD-MM-YYYY HH:MM"
                cell.number_format = number_format

    # =========================
    # AUTO WIDTH
    # =========================
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    # =========================
    # SHEET FEATURES & RESPONSE
    # =========================
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    xlsx_file = BytesIO()
    wb.save(xlsx_file)

    # Populate standard Frappe file download response headers
    frappe.response["filename"] = f"{report_name}.xlsx"
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"