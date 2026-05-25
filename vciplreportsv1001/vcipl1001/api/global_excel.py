import frappe
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from frappe.utils.xlsxutils import handle_html

def custom_make_xlsx(data, sheet_name="Sheet 1", wb=None, column_widths=None):
    """
    Global interceptor for Frappe Excel generation.
    Reads styling from 'Excel Export Settings' DocType.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]

    # =========================
    # FETCH SETTINGS FROM UI
    # =========================
    # Fallbacks in case the DocType hasn't been saved yet
    bg_color_hex = "1F4E78" 
    text_color_hex = "FFFFFF"
    enable_zebra = 1

    try:
        settings = frappe.get_single("Excel Export Settings")
        # openpyxl needs colors without the '#' symbol
        if settings.header_bg_color:
            bg_color_hex = settings.header_bg_color.replace("#", "")
        if settings.header_text_color:
            text_color_hex = settings.header_text_color.replace("#", "")
        enable_zebra = settings.enable_zebra
    except Exception:
        pass # If DocType doesn't exist yet, just use fallbacks

    # =========================
    # DESIGN STYLES
    # =========================
    header_fill = PatternFill(start_color=bg_color_hex, end_color=bg_color_hex, fill_type="solid")
    header_font = Font(bold=True, name="Calibri", size=12, color=text_color_hex)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Very light grey for alternate rows (Zebra Striping)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="A6A6A6"), right=Side(style="thin", color="A6A6A6"),
        top=Side(style="thin", color="A6A6A6"), bottom=Side(style="thin", color="A6A6A6"),
    )

    # =========================
    # POPULATE & STYLE DATA
    # =========================
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            
            if isinstance(value, str):
                if value.startswith("="):
                    value = f"'{value}"
                elif sheet_name not in ["Data Import Template", "Data Export"]:
                    value = handle_html(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Apply Styles based on row
            if row_idx == 1:
                # Row 1 is Header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                # Data Rows - Apply Zebra Striping
                if enable_zebra and (row_idx % 2 == 0):
                    cell.fill = zebra_fill

    # =========================
    # AUTO-FIT COLUMNS & FREEZE
    # =========================
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"
    
    # Add Auto-Filters to the header row
    ws.auto_filter.ref = ws.dimensions

    # =========================
    # RETURN BYTES
    # =========================
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file